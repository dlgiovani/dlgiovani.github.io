from tkinter import *
from itertools import count
import re
import tkinter
from tkinter.filedialog import *
from urllib.request import OpenerDirector
import openpyxl

from setuptools import Command

def getFile(fileList):
    path = fileList
    xcel_i_row = 1
    if path != "":
        wb = openpyxl.Workbook()
        sheet = wb.active
        outputFileName = path + "-word rank by Giovani Drosda Lima.txt"
        outputFile = open(outputFileName, "w")

        fileSrc = open(path, "rt")
        file = fileSrc.read()

        noDotsText = re.sub(r"[^a-zA-Z0-9-\s]", "", file)
        #noDotsText = re.sub('["]', "", file)

        listOfWords = re.split("\s", noDotsText)
        listOfWordsCounted = list()
        listOfWordsWithCountDict = {}

        for eachWord in listOfWords:
            if eachWord not in listOfWordsCounted:
                x = listOfWords.count(eachWord)

                sheet.cell(row = xcel_i_row, column = 1).value = eachWord #for the excel file
                sheet.cell(row = xcel_i_row, column = 2).value = x

                listOfWordsCounted.append(eachWord) #for the txt file
                listOfWordsWithCountDict[eachWord] = x

                xcel_i_row += 1

        wb.save(path + " word rank by Giovani Drosda Lima.xlsx")

        sortedList = sorted(listOfWordsWithCountDict.items(), key=lambda x: (1-x[1], x[0]))
        for item in sortedList:
            #print(item, "\n")
            line = str(item) + "\n"
            outputFile.write(line)

            return(outputFile)
        
        fileSrc.close()
        outputFile.close()


